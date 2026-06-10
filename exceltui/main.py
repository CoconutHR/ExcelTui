import re
import json
import math
from rich.console import Console
from rich.panel import Panel
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn
from rich.prompt import Prompt, Confirm
from rich.table import Table
from rich import box
import os
import sys
import shutil
import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import numbers as openpyxl_numbers

console = Console()

# openpyxl 不允许写入的控制字符（会触发 IllegalCharacterError）
# 规则参考 openpyxl.cell.cell.ILLEGAL_CHARACTERS_RE
_ILLEGAL_EXCEL_CHAR_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def _is_empty_value(value):
    """判断值是否为空（None / NaN / 空字符串 / 'nan'）。替代 pd.isna。"""
    if value is None:
        return True
    if isinstance(value, float):
        # NaN check without pandas
        return value != value
    if isinstance(value, str) and value.strip().lower() == "nan":
        return True
    return False


def sanitize_excel_cell_value(value):
    """清理 Excel 单元格中 openpyxl 不允许的字符"""
    if value is None:
        return value

    if _is_empty_value(value):
        return value

    text = str(value)
    return _ILLEGAL_EXCEL_CHAR_RE.sub("", text)


def _force_text_format_worksheet(worksheet):
    """将工作表中所有单元格设置为文本格式，并将非空值强制转为字符串，防止科学计数法"""
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                   min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.number_format = "@"
            if cell.value is not None:
                if not _is_empty_value(cell.value):
                    cell.value = sanitize_excel_cell_value(cell.value)


def _read_xlsx_to_rows(filepath):
    """用 openpyxl 读取 xlsx，返回 (headers: list[str], data_rows: list[list[str]])。
    所有值统一转为字符串或 None。"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    # 第一行为表头
    raw_headers = next(rows_iter, None)
    if raw_headers is None:
        wb.close()
        return [], []
    headers = [str(h) if h is not None else f"Column_{i}" for i, h in enumerate(raw_headers)]

    data_rows = []
    for raw_row in rows_iter:
        row = []
        for val in raw_row:
            if val is None:
                row.append(None)
            else:
                row.append(str(val))
        data_rows.append(row)
    wb.close()
    return headers, data_rows


def _write_rows_to_xlsx(filepath, headers, data_rows):
    """用 openpyxl 将 headers + data_rows 写入 xlsx，所有单元格设为文本格式。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    # 写入表头
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.number_format = "@"
    # 写入数据行
    for ri, row in enumerate(data_rows, start=2):
        for ci, val in enumerate(row, start=1):
            cleaned = sanitize_excel_cell_value(val) if val is not None else None
            cell = ws.cell(row=ri, column=ci, value=cleaned)
            cell.number_format = "@"
    wb.save(filepath)
    wb.close()

class SimpleTable:
    """轻量级表格数据结构，替代 pandas DataFrame，支持 table.columns / table[col] / len(table) 等操作"""

    def __init__(self, headers, data_rows):
        """
        headers: list[str] — 列名列表
        data_rows: list[list] — 每行一个列表，长度应与 headers 一致
        """
        self.columns = list(headers)
        self._data = [list(row) for row in data_rows]
        # 补齐行宽不一致的情况
        width = len(self.columns)
        for row in self._data:
            while len(row) < width:
                row.append(None)

    def __len__(self):
        return len(self._data)

    def __getitem__(self, col_name):
        """table[col_name] → 返回该列所有值的列表"""
        ci = self.columns.index(col_name)
        return [row[ci] for row in self._data]

    def __setitem__(self, col_name, values):
        """table[col_name] = values → 设置/新增列"""
        if col_name in self.columns:
            ci = self.columns.index(col_name)
            for ri, val in enumerate(values):
                self._data[ri][ci] = val
        else:
            self.columns.append(col_name)
            for ri, val in enumerate(values):
                self._data[ri].append(val)

    def copy(self):
        return SimpleTable(self.columns[:], [row[:] for row in self._data])

    def get_row(self, index):
        """返回第 index 行（0-based）的值列表"""
        return self._data[index]

    def get_rows(self):
        """返回所有数据行"""
        return self._data

    def apply_sanitize(self):
        """对所有单元格执行 sanitize_excel_cell_value"""
        for row in self._data:
            for ci in range(len(row)):
                row[ci] = sanitize_excel_cell_value(row[ci])
        return self

    @staticmethod
    def from_xlsx(filepath):
        """从 xlsx 文件读取"""
        headers, data_rows = _read_xlsx_to_rows(filepath)
        return SimpleTable(headers, data_rows)

    def to_xlsx(self, filepath):
        """写入 xlsx 文件（带文本格式和字符清理）"""
        _write_rows_to_xlsx(filepath, self.columns, self._data)


class DataFormat:
    """数据格式枚举"""
    JSON = "json"
    KEY_VALUE = "key_value"
    PLAIN_TEXT = "plain_text"

def detect_format(text):
    """检测文本格式"""
    if _is_empty_value(text):
        return None
    
    text = str(text).strip()
    
    # 尝试解析为JSON
    try:
        json.loads(text)
        return DataFormat.JSON
    except (json.JSONDecodeError, ValueError, TypeError):
        pass
    
    # 检测是否为键值对格式 {key:value}
    if re.search(r'\{[^:}]+:.*?\}', text):
        return DataFormat.KEY_VALUE
    
    # 否则为纯文本
    return DataFormat.PLAIN_TEXT

def extract_keys_from_json(text):
    """从JSON格式提取键"""
    try:
        data = json.loads(str(text))
        if isinstance(data, dict):
            return list(data.keys())
    except (json.JSONDecodeError, ValueError, TypeError):
        pass
    return []

def extract_keys_from_key_value(text):
    """从键值对格式提取键 - 修复版本"""
    if _is_empty_value(text):
        return []
    
    text = str(text)
    keys = []
    
    i = 0
    while i < len(text):
        if text[i] == '{':
            colon_pos = text.find(':', i)
            if colon_pos == -1:
                i += 1
                continue
            
            brace_count = 1
            j = i + 1
            right_brace_pos = -1
            
            while j < len(text) and brace_count > 0:
                if text[j] == '{':
                    brace_count += 1
                elif text[j] == '}':
                    brace_count -= 1
                    if brace_count == 0:
                        right_brace_pos = j
                        break
                j += 1
            
            if right_brace_pos != -1 and colon_pos < right_brace_pos:
                key = text[i+1:colon_pos].strip()
                if key:
                    keys.append(key)
                i = right_brace_pos + 1
            else:
                i += 1
        else:
            i += 1
    
    return keys

def scan_column_format(df, column_name):
    """扫描列的格式分布"""
    format_counts = {
        DataFormat.JSON: 0,
        DataFormat.KEY_VALUE: 0,
        DataFormat.PLAIN_TEXT: 0,
        None: 0
    }
    
    all_keys = set()
    
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        console=console
    ) as progress:
        task = progress.add_task("[cyan]正在扫描数据格式...", total=len(df))
        
        for idx, value in enumerate(df[column_name]):
            fmt = detect_format(value)
            format_counts[fmt] = format_counts.get(fmt, 0) + 1
            
            if fmt == DataFormat.JSON:
                keys = extract_keys_from_json(value)
                all_keys.update(keys)
            elif fmt == DataFormat.KEY_VALUE:
                keys = extract_keys_from_key_value(value)
                all_keys.update(keys)
            
            progress.update(task, advance=1)
    
    return format_counts, sorted(list(all_keys))

def extract_value_from_json(text, key):
    """从JSON中提取值"""
    try:
        data = json.loads(str(text))
        if isinstance(data, dict) and key in data:
            return data[key]
    except (json.JSONDecodeError, ValueError, TypeError):
        pass
    return ""

def extract_value_from_key_value(text, key):
    """从键值对中提取值"""
    if _is_empty_value(text):
        return ""
    
    text = str(text)
    escaped_key = re.escape(key)
    pattern = rf'\{{{escaped_key}:'
    match = re.search(pattern, text)
    
    if not match:
        return ""
    
    start_pos = match.end()
    brace_count = 1
    i = start_pos
    
    while i < len(text) and brace_count > 0:
        if text[i] == '{':
            brace_count += 1
        elif text[i] == '}':
            brace_count -= 1
            if brace_count == 0:
                value = text[start_pos:i].strip()
                return value
        i += 1
    
    return ""

def extract_key_value_pair(text, key):
    """从键值对中提取完整的键值对"""
    if _is_empty_value(text):
        return ""
    
    text = str(text)
    escaped_key = re.escape(key)
    pattern = rf'\{{{escaped_key}:'
    match = re.search(pattern, text)
    
    if not match:
        return ""
    
    start_pos = match.start()
    brace_count = 1
    i = match.end()
    
    while i < len(text) and brace_count > 0:
        if text[i] == '{':
            brace_count += 1
        elif text[i] == '}':
            brace_count -= 1
            if brace_count == 0:
                return text[start_pos:i+1]
        i += 1
    
    return ""

def remove_key_from_json(text, key):
    """从JSON中删除键"""
    try:
        data = json.loads(str(text))
        if isinstance(data, dict) and key in data:
            del data[key]
            return json.dumps(data, ensure_ascii=False)
    except (json.JSONDecodeError, ValueError, TypeError):
        pass
    return text

def remove_key_from_key_value(text, key):
    """从键值对中删除键"""
    if _is_empty_value(text):
        return ""
    
    text = str(text)
    key_value_pair = extract_key_value_pair(text, key)
    if key_value_pair:
        text = text.replace(key_value_pair + '\n', '')
        text = text.replace('\n' + key_value_pair, '')
        text = text.replace(key_value_pair, '')
    
    return text.strip()

def process_escape_chars(text):
    """处理转义字符"""
    if _is_empty_value(text):
        return ""
    
    text = str(text)
    text = text.replace('\\n', '\n')
    text = text.replace('\\t', '\t')
    text = text.replace('\\r', '\r')
    text = text.replace('\\"', '"')
    text = text.replace("\\'", "'")
    text = text.replace('\\\\', '\\')
    
    return text

def show_format_distribution(format_counts):
    """显示格式分布"""
    table = Table(title="数据格式分布", box=box.ROUNDED)
    table.add_column("格式类型", style="cyan")
    table.add_column("数量", style="yellow", justify="right")
    table.add_column("占比", style="green", justify="right")
    
    total = sum(format_counts.values())
    
    format_names = {
        DataFormat.JSON: "JSON格式",
        DataFormat.KEY_VALUE: "键值对格式 {key:value}",
        DataFormat.PLAIN_TEXT: "纯文本",
        None: "空值"
    }
    
    for fmt, count in format_counts.items():
        if count > 0:
            percentage = f"{count / total * 100:.1f}%"
            table.add_row(format_names.get(fmt, "未知"), str(count), percentage)
    
    console.print(table)

def show_keys_table(keys, page_size=20):
    """显示键列表"""
    if not keys:
        console.print("[yellow]未检测到任何键[/yellow]")
        return
    
    total_pages = (len(keys) + page_size - 1) // page_size
    current_page = 0
    
    while True:
        console.clear()
        start_idx = current_page * page_size
        end_idx = min(start_idx + page_size, len(keys))
        page_keys = keys[start_idx:end_idx]
        
        table = Table(
            title=f"检测到的键 (共 {len(keys)} 个) - 第 {current_page + 1}/{total_pages} 页",
            box=box.ROUNDED
        )
        table.add_column("序号", style="cyan", justify="center", width=8)
        table.add_column("键名", style="green")
        
        for i, key in enumerate(page_keys, start=start_idx + 1):
            table.add_row(str(i), key)
        
        console.print(table)
        
        if total_pages > 1:
            console.print(f"\n[dim]提示: 输入 'n' 下一页, 'p' 上一页, 'q' 继续[/dim]")
            nav = Prompt.ask("导航", choices=["n", "p", "q"], default="q")
            
            if nav == "n" and current_page < total_pages - 1:
                current_page += 1
            elif nav == "p" and current_page > 0:
                current_page -= 1
            elif nav == "q":
                break
        else:
            break

def select_keys(all_keys):
    """让用户选择要提取的键"""
    if not all_keys:
        console.print("[yellow]没有可选择的键[/yellow]")
        return None
    
    console.print("\n[bold cyan]请选择要提取的键:[/bold cyan]")
    console.print("  [dim]输入格式: 可以输入序号(如: 1,3,5)或键名[/dim]")
    console.print("  [dim]输入 'all' 提取所有键[/dim]")
    console.print("  [dim]输入 'q' 返回上一步[/dim]\n")
    
    while True:
        selection = Prompt.ask("请输入选择")
        
        if selection.lower() == 'q':
            return None
        
        if selection.lower() == 'all':
            return all_keys
        
        selected_keys = []
        parts = [p.strip() for p in selection.split(',')]
        
        for part in parts:
            try:
                idx = int(part) - 1
                if 0 <= idx < len(all_keys):
                    selected_keys.append(all_keys[idx])
                else:
                    console.print(f"[yellow]⚠ 序号 {part} 超出范围[/yellow]")
            except ValueError:
                if part in all_keys:
                    selected_keys.append(part)
                else:
                    console.print(f"[yellow]⚠ 键名 '{part}' 不存在[/yellow]")
        
        if selected_keys:
            console.print(f"\n[green]✓ 已选择 {len(selected_keys)} 个键:[/green]")
            for key in selected_keys:
                console.print(f"  • {key}")
            
            if Confirm.ask("\n确认选择?", default=True):
                return selected_keys
        else:
            console.print("[red]✗ 没有选择任何有效的键[/red]")

def get_processing_options(has_keys):
    """获取处理选项"""
    options = {}
    
    console.print("\n[bold cyan]处理选项配置:[/bold cyan]")
    
    options['process_escape'] = Confirm.ask(
        "是否处理转义字符(如 \\n 转为换行)?",
        default=True
    )
    
    options['strip_values'] = Confirm.ask(
        "是否去除值的首尾空白字符(.strip())?",
        default=True
    )
    
    if has_keys:
        console.print("\n[bold]写入格式:[/bold]")
        console.print("  [cyan]1.[/cyan] 只写入值(键作为列名)")
        console.print("  [cyan]2.[/cyan] 写入完整键值对 {键:值}")
        
        write_format = Prompt.ask(
            "请选择",
            choices=["1", "2"],
            default="1"
        )
        options['write_key_value'] = (write_format == "2")
    else:
        options['write_key_value'] = False
    
    return options

def process_value(value, key, options, data_format):
    """根据选项处理值"""
    if _is_empty_value(value) or value == '':
        return ""
    
    value = str(value)
    
    if options.get('process_escape', False):
        value = process_escape_chars(value)
    
    if options.get('strip_values', False):
        value = value.strip()
    
    if options.get('write_key_value', False) and key:
        return f"{{{key}:{value}}}"
    
    return value

def show_operation_menu(current_column=None):
    """显示操作菜单"""
    console.print("\n[bold cyan]═══════════════════════════════════════[/bold cyan]")
    if current_column:
        console.print(f"[bold]当前处理列: [green]{current_column}[/green][/bold]")
    console.print("[bold]请选择操作:[/bold]")
    console.print("  [cyan]1.[/cyan] 提取键值对到新列")
    console.print("  [cyan]2.[/cyan] 删除指定键值对")
    console.print("  [cyan]3.[/cyan] 处理纯文本列")
    console.print("  [cyan]4.[/cyan] 切换到其他列")
    console.print("  [cyan]5.[/cyan] 保存并退出")
    console.print("  [cyan]6.[/cyan] 放弃修改并退出")
    
    return Prompt.ask("\n请输入选项", choices=["1", "2", "3", "4", "5", "6"], default="1")

def show_columns_menu(df, current_column=None):
    """显示列选择菜单"""
    console.print("\n[bold cyan]可用的列:[/bold cyan]")
    
    table = Table(box=box.ROUNDED)
    table.add_column("序号", style="cyan", justify="center", width=8)
    table.add_column("列名", style="green")
    table.add_column("状态", style="yellow", width=10)
    
    for idx, col in enumerate(df.columns):
        status = "当前列" if col == current_column else ""
        table.add_row(str(idx + 1), col, status)
    
    console.print(table)
    
    col_choice = Prompt.ask(
        f"\n请选择要处理的列 (1-{len(df.columns)}) 或输入 'q' 返回",
        default="q"
    )
    
    if col_choice.lower() == 'q':
        return None
    
    try:
        idx = int(col_choice) - 1
        if 0 <= idx < len(df.columns):
            return df.columns[idx]
        else:
            console.print("[red]✗ 无效的列编号[/red]")
            return None
    except ValueError:
        console.print("[red]✗ 请输入有效的数字[/red]")
        return None

def extract_keys_operation(df, column_name, all_keys, format_counts):
    """提取键值对操作"""
    console.print("\n[bold cyan]提取键值对到新列[/bold cyan]")

    show_keys_table(all_keys)

    selected_keys = select_keys(all_keys)
    if selected_keys is None:
        return df, False

    # 默认使用“键名作为列名”，同时支持用户自定义列名
    # 若自定义列名与已有列重名，后续直接覆盖
    target_columns = {}
    custom_names = Confirm.ask("是否为提取结果自定义列名? (默认直接使用键名)", default=False)
    for key in selected_keys:
        if custom_names:
            col_name = Prompt.ask(f"键 [{key}] 的列名", default=key).strip()
            target_columns[key] = col_name if col_name else key
        else:
            target_columns[key] = key

    options = get_processing_options(has_keys=True)

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        console=console
    ) as progress:
        task = progress.add_task(
            f"[cyan]正在提取 {len(selected_keys)} 个键值对...",
            total=len(selected_keys)
        )

        for key in selected_keys:
            values = []

            for idx, cell_value in enumerate(df[column_name]):
                fmt = detect_format(cell_value)

                if fmt == DataFormat.JSON:
                    raw_value = extract_value_from_json(cell_value, key)
                elif fmt == DataFormat.KEY_VALUE:
                    if options.get('write_key_value', False):
                        raw_value = extract_key_value_pair(cell_value, key)
                    else:
                        raw_value = extract_value_from_key_value(cell_value, key)
                else:
                    raw_value = ""

                processed = process_value(raw_value, key if not options.get('write_key_value', False) else None, options, fmt)
                values.append(processed)

            target_col = target_columns[key]
            df[target_col] = values
            progress.update(task, advance=1)

    console.print(f"[green]✓ 成功提取 {len(selected_keys)} 个键到新列[/green]")
    return df, True

def delete_keys_operation(df, column_name, all_keys):
    """删除键值对操作"""
    console.print("\n[bold cyan]删除指定键值对[/bold cyan]")
    
    show_keys_table(all_keys)
    
    selected_keys = select_keys(all_keys)
    if selected_keys is None:
        return df, False
    
    if not Confirm.ask(f"\n确认从源列中删除这 {len(selected_keys)} 个键值对?", default=False):
        return df, False
    
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        console=console
    ) as progress:
        task = progress.add_task(
            f"[cyan]正在删除 {len(selected_keys)} 个键值对...",
            total=len(df)
        )
        
        new_values = []
        for cell_value in df[column_name]:
            fmt = detect_format(cell_value)
            current_value = cell_value
            
            for key in selected_keys:
                if fmt == DataFormat.JSON:
                    current_value = remove_key_from_json(current_value, key)
                elif fmt == DataFormat.KEY_VALUE:
                    current_value = remove_key_from_key_value(current_value, key)
            
            new_values.append(current_value)
            progress.update(task, advance=1)
        
        df[column_name] = new_values
    
    console.print(f"[green]✓ 成功删除 {len(selected_keys)} 个键值对[/green]")
    return df, True

def process_plain_text_operation(df, column_name):
    """处理纯文本操作"""
    console.print("\n[bold cyan]处理纯文本列[/bold cyan]")
    
    options = get_processing_options(has_keys=False)
    
    new_column_name = Prompt.ask(
        "请输入新列名",
        default=f"{column_name}_处理后"
    )
    
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        console=console
    ) as progress:
        task = progress.add_task("[cyan]正在处理纯文本...", total=len(df))
        
        processed_values = []
        for value in df[column_name]:
            processed = process_value(value, None, options, DataFormat.PLAIN_TEXT)
            processed_values.append(processed)
            progress.update(task, advance=1)
        
        df[new_column_name] = processed_values
    
    console.print(f"[green]✓ 成功处理纯文本到新列: {new_column_name}[/green]")
    return df, True

def process_column_interactive(df, column_name):
    """交互式处理单个列"""
    format_counts, all_keys = scan_column_format(df, column_name)
    
    show_format_distribution(format_counts)
    
    main_format = max(format_counts.items(), key=lambda x: x[1] if x[0] is not None else 0)[0]
    
    if main_format == DataFormat.PLAIN_TEXT and not all_keys:
        console.print("\n[yellow]检测到主要为纯文本格式,无结构化键值对[/yellow]")
    
    modified = False
    
    while True:
        operation = show_operation_menu(column_name)
        
        if operation == "1":
            if not all_keys:
                console.print("[yellow]当前列没有检测到键值对[/yellow]")
                continue
            
            df, success = extract_keys_operation(df, column_name, all_keys, format_counts)
            if success:
                modified = True
                format_counts, all_keys = scan_column_format(df, column_name)
        
        elif operation == "2":
            if not all_keys:
                console.print("[yellow]当前列没有检测到键值对[/yellow]")
                continue
            
            df, success = delete_keys_operation(df, column_name, all_keys)
            if success:
                modified = True
                format_counts, all_keys = scan_column_format(df, column_name)
        
        elif operation == "3":
            df, success = process_plain_text_operation(df, column_name)
            if success:
                modified = True
        
        elif operation == "4":
            return df, modified, "switch_column"
        
        elif operation == "5":
            return df, modified, "save"
        
        elif operation == "6":
            return df, modified, "cancel"

def process_excel_interactive(file_path):
    """交互式处理Excel文件"""
    try:
        console.print(f"[cyan]正在读取文件: {file_path}[/cyan]")
        df = SimpleTable.from_xlsx(file_path)
        console.print(f"[green]✓[/green] 成功读取 {len(df)} 行数据")
        
        current_column = show_columns_menu(df)
        if current_column is None:
            return False, "用户取消操作", 0
        
        console.print(f"[cyan]处理列: {current_column}[/cyan]\n")
        
        global_modified = False
        
        while True:
            df, modified, action = process_column_interactive(df, current_column)
            
            if modified:
                global_modified = True
            
            if action == "switch_column":
                new_column = show_columns_menu(df, current_column)
                if new_column is None:
                    continue
                current_column = new_column
                console.print(f"[cyan]切换到列: {current_column}[/cyan]\n")
            
            elif action == "save":
                if not global_modified:
                    console.print("[yellow]没有进行任何修改[/yellow]")
                    if not Confirm.ask("确认退出?", default=True):
                        continue
                    return False, "未进行修改", 0
                
                base, _ext = os.path.splitext(file_path)
                output_file = base + "_处理后.xlsx"
                
                if os.path.exists(output_file):
                    if not Confirm.ask(f"文件 {output_file} 已存在,是否覆盖?", default=True):
                        new_name = Prompt.ask("请输入新的文件名")
                        if not new_name.endswith('.xlsx'):
                            new_name += '.xlsx'
                        output_file = new_name
                
                console.print("\n[cyan]正在保存文件...[/cyan]")

                df.apply_sanitize()
                df.to_xlsx(output_file)
                
                return True, output_file, len(df)
            
            elif action == "cancel":
                if global_modified:
                    if not Confirm.ask("确认放弃所有修改?", default=False):
                        continue
                
                return False, "用户取消操作", 0
    
    except Exception as e:
        import traceback
        return False, f"{str(e)}\n{traceback.format_exc()}", 0

# ────────────────────────────────────────────────────────────────────────────────
# 通用工具函数（用于新功能）
# ────────────────────────────────────────────────────────────────────────────────

def index_to_col_letter(index):
    """将0-based索引转换为列字母（如0->A，25->Z，26->AA）"""
    result = ""
    index += 1
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        result = chr(ord('A') + remainder) + result
    return result


def col_letter_to_index(col_letter):
    """将列字母转换为0-based索引"""
    col_letter = col_letter.upper()
    index = 0
    for char in col_letter:
        index = index * 26 + (ord(char) - ord('A') + 1)
    return index - 1


def is_valid_col_letter(s):
    """判断字符串是否为合法的列字母（仅含A-Z）"""
    return len(s) > 0 and all('A' <= c <= 'Z' for c in s.upper())


def get_file_size_str(filepath):
    """返回人类可读的文件大小字符串"""
    size = os.path.getsize(filepath)
    if size < 1024:
        return f"{size} B"
    elif size < 1024 * 1024:
        return f"{size / 1024:.1f} KB"
    else:
        return f"{size / 1024 / 1024:.1f} MB"


def get_xlsx_files():
    """获取当前工作目录下的所有xlsx文件"""
    current_dir = Path.cwd()
    xlsx_files = sorted([f for f in current_dir.glob('*.xlsx') if not f.name.startswith('~$')])
    return xlsx_files, str(current_dir)


def wrap_text(text, max_width):
    """智能文本换行，支持长文件名"""
    if len(text) <= max_width:
        return text
    
    # 对于长文件名，智能截断并显示省略号
    if len(text) > max_width:
        # 保留前后部分，中间用省略号
        half = (max_width - 3) // 2
        return text[:half] + "..." + text[-(max_width - half - 3):]
    return text


def get_safe_filename(filepath_or_name):
    """获取安全的文件名显示（只显示文件名，不包含路径）"""
    if isinstance(filepath_or_name, Path):
        return filepath_or_name.name
    elif isinstance(filepath_or_name, str):
        return os.path.basename(filepath_or_name)
    return str(filepath_or_name)


def display_file_list_tui(xlsx_files, script_dir, title="📂 当前文件夹中的 Excel 文件"):
    """以表格形式展示 xlsx 文件列表，优化长文件名显示"""
    # 获取终端宽度，默认80
    try:
        terminal_width = shutil.get_terminal_size().columns
    except (ValueError, OSError):
        terminal_width = 80
    
    # 计算文件名列的宽度（留出编号和大小的空间）
    col_widths = {
        '编号': 6,
        '大小': 12,
        '间隔': 4  # 列之间的间隔
    }
    filename_width = terminal_width - col_widths['编号'] - col_widths['大小'] - col_widths['间隔']
    filename_width = max(filename_width, 30)  # 至少30个字符宽
    
    table = Table(
        title=title,
        show_header=True,
        header_style="bold cyan",
        box=box.ROUNDED,
        show_lines=False,
        padding=(0, 1)
    )
    table.add_column("编号", style="bold yellow", justify="center", width=6)
    table.add_column("文件名", style="white", width=filename_width, overflow="fold")
    table.add_column("文件大小", style="dim", justify="right", width=10)

    for idx, file_item in enumerate(xlsx_files, start=1):
        # 统一获取文件名
        filename_str = get_safe_filename(file_item)
        
        # 构建完整路径用于计算文件大小
        filepath = os.path.join(script_dir, filename_str)
        
        # 如果文件名过长，进行智能处理
        if len(filename_str) > filename_width:
            display_name = wrap_text(filename_str, filename_width)
        else:
            display_name = filename_str
            
        table.add_row(str(idx), display_name, get_file_size_str(filepath))

    console.print(table)
    
    # 如果有长文件名被截断，显示提示
    long_files_count = sum(1 for f in xlsx_files if len(get_safe_filename(f)) > filename_width)
    if long_files_count > 0:
        console.print(f"\n[dim]💡 提示: 共有 {long_files_count} 个文件名超过显示宽度，已自动优化显示[/dim]")


def select_file_tui(xlsx_files, prompt_text="请选择要处理的文件编号或输入 'q' 返回"):
    """通用文件选择，返回 0-based 索引；用户输入 q 或直接按回车时返回 None"""
    while True:
        user_input = Prompt.ask(f"\n[bold green]{prompt_text}[/bold green]", default="q").strip()

        if user_input.lower() == "q":
            console.print("[dim]已返回上级菜单。[/dim]")
            return None

        if not user_input.isdigit():
            console.print("[bold red]❌ 请输入有效的数字编号！[/bold red]")
            continue

        idx = int(user_input)
        if idx < 1 or idx > len(xlsx_files):
            console.print(
                f"[bold red]❌ 编号超出范围，请输入 1 到 {len(xlsx_files)} 之间的数字！[/bold red]"
            )
            continue

        return idx - 1


def select_column_interactively(ws, prompt_title="请输入列号（如 A、B、P、AA 等），输入 q 返回"):
    """展示工作表列预览，提示用户输入列字母，返回 (col_letter, col_index_0based)；用户输入 q 时返回 None"""
    max_col = ws.max_column

    # 读取表头行
    headers = []
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        headers.append(str(cell.value) if cell.value is not None else "")

    # 展示列预览（最多 52 列）
    col_table = Table(title="📋 文件列信息预览", show_header=True, header_style="bold cyan", box=box.ROUNDED)
    col_table.add_column("列字母", style="bold yellow", justify="center")
    col_table.add_column("列名（表头）", style="white")

    preview_count = min(max_col, 52)
    for i in range(preview_count):
        col_table.add_row(index_to_col_letter(i), headers[i])
    if max_col > preview_count:
        col_table.add_row("...", f"（共 {max_col} 列）")

    console.print(col_table)

    while True:
        user_input = Prompt.ask(f"\n[bold green]{prompt_title}[/bold green]").strip()

        if user_input.lower() == "q":
            console.print("[dim]已返回上级菜单。[/dim]")
            return None

        if not user_input:
            console.print("[bold red]❌ 输入不能为空，请重新输入！[/bold red]")
            continue

        # 自动处理小写
        if user_input != user_input.upper():
            console.print(
                f"[bold yellow]⚠️  检测到小写字母，已自动转换为大写："
                f"[bold white]{user_input.upper()}[/bold white][/bold yellow]"
            )
            user_input = user_input.upper()

        # 非法字符检测
        if not is_valid_col_letter(user_input):
            console.print(
                "[bold red]❌ 列号包含非法字符，只能输入字母（A-Z），请重新输入！[/bold red]"
            )
            continue

        col_index = col_letter_to_index(user_input)

        # 超出范围检测
        if col_index >= max_col:
            console.print(
                f"[bold red]❌ 列号 {user_input} 超出文件范围"
                f"（共 {max_col} 列，最大列为 {index_to_col_letter(max_col - 1)}），请重新输入！[/bold red]"
            )
            continue

        return user_input, col_index


def select_column_from_df(df, prompt_title="请输入列号（如 A、B、P、AA 等），输入 q 返回"):
    """从DataFrame展示列预览，提示用户输入列字母，返回 (col_letter, col_index_0based)；用户输入 q 时返回 None"""
    total_cols = len(df.columns)

    # 展示列预览（最多 52 列）
    col_table = Table(title="📋 文件列信息预览", show_header=True, header_style="bold cyan", box=box.ROUNDED)
    col_table.add_column("列字母", style="bold yellow", justify="center")
    col_table.add_column("列名（表头）", style="white")

    preview_count = min(total_cols, 52)
    for i in range(preview_count):
        col_table.add_row(index_to_col_letter(i), str(df.columns[i]))
    if total_cols > preview_count:
        col_table.add_row("...", f"（共 {total_cols} 列）")

    console.print(col_table)

    while True:
        user_input = Prompt.ask(f"\n[bold green]{prompt_title}[/bold green]").strip()

        if user_input.lower() == "q":
            console.print("[dim]已返回上级菜单。[/dim]")
            return None

        if not user_input:
            console.print("[bold red]❌ 输入不能为空，请重新输入！[/bold red]")
            continue

        # 自动处理小写
        if user_input != user_input.upper():
            console.print(
                f"[bold yellow]⚠️  检测到小写字母，已自动转换为大写："
                f"[bold white]{user_input.upper()}[/bold white][/bold yellow]"
            )
            user_input = user_input.upper()

        # 非法字符检测
        if not is_valid_col_letter(user_input):
            console.print(
                "[bold red]❌ 列号包含非法字符，只能输入字母（A-Z），请重新输入！[/bold red]"
            )
            continue

        col_index = col_letter_to_index(user_input)

        # 超出范围检测
        if col_index >= total_cols:
            console.print(
                f"[bold red]❌ 列号 {user_input} 超出文件范围"
                f"（共 {total_cols} 列，最大列为 {index_to_col_letter(total_cols - 1)}），请重新输入！[/bold red]"
            )
            continue

        return user_input, col_index


# ────────────────────────────────────────────────────────────────────────────────
# 功能1：去除xlsx指定列的重复项
# ────────────────────────────────────────────────────────────────────────────────

def deduplicate_by_column():
    """根据指定列去重，保留首次出现的行"""
    try:
        console.print(Panel.fit(
            "[bold cyan]🔁 Excel 去重处理工具[/bold cyan]\n[dim]根据指定列去重，保留首次出现的行[/dim]",
            border_style="cyan",
            padding=(1, 4)
        ))
        console.print()  # 空行美化

        xlsx_files, script_dir = get_xlsx_files()

        if not xlsx_files:
            console.print(Panel(
                "[bold red]❌ 当前文件夹中没有找到任何 .xlsx 文件！\n请将需要处理的 Excel 文件放在本程序所在文件夹中。[/bold red]",
                border_style="red"
            ))
            Prompt.ask("\n[dim]按回车返回上级菜单[/dim]", default="q")
            return False

        # 显示文件列表
        display_file_list_tui(xlsx_files, script_dir)

        # 用户选择文件
        file_idx = select_file_tui(xlsx_files)
        if file_idx is None:
            return False
        selected_file = get_safe_filename(xlsx_files[file_idx])
        filepath = os.path.join(script_dir, selected_file)

        console.print(f"\n[bold green]✅ 已选择文件：[bold white]{selected_file}[/bold white][/bold green]")

        # 读取文件
        with console.status("[bold cyan]正在读取文件...[/bold cyan]", spinner="dots"):
            tbl = SimpleTable.from_xlsx(filepath)

        # 用户选择列
        result = select_column_from_df(tbl)
        if result is None:
            return False
        col_letter, col_index = result

        # 确认操作
        console.print()
        console.print(Panel(
            f"[bold]操作确认[/bold]\n\n"
            f"  📄 处理文件：[cyan]{selected_file}[/cyan]\n"
            f"  🔑 去重列：  [cyan]{col_letter} 列[/cyan]（{tbl.columns[col_index]}）\n"
            f"  📁 输出目录：[cyan]{script_dir}[/cyan]",
            border_style="yellow",
            title="[yellow]请确认[/yellow]"
        ))

        confirm = Prompt.ask(
            "[bold yellow]确认开始处理？[/bold yellow]",
            choices=["y", "n"],
            default="y"
        )

        if confirm.lower() != 'y':
            console.print("[dim]已取消操作。[/dim]")
            return False

        # 执行去重
        console.print()
        original_count = len(tbl)
        console.print(f"[green]✅ 文件读取完成，共 [bold]{original_count}[/bold] 行数据[/green]")

        with console.status(f"[bold cyan]正在根据 {col_letter} 列去重...[/bold cyan]", spinner="dots"):
            seen = set()
            dedup_rows = []
            for row in tbl.get_rows():
                key = row[col_index]
                key_str = str(key).strip() if key is not None else ""
                if key_str not in seen:
                    seen.add(key_str)
                    dedup_rows.append(row)

        dedup_count = len(dedup_rows)
        removed_count = original_count - dedup_count

        # 生成输出文件名
        base_name = os.path.splitext(selected_file)[0]
        output_filename = f"{base_name}_去重_{col_letter}列.xlsx"
        output_path = os.path.join(script_dir, output_filename)

        with console.status(f"[bold cyan]正在保存文件...[/bold cyan]", spinner="dots"):
            _write_rows_to_xlsx(output_path, tbl.columns, dedup_rows)

        # 显示结果
        result_table = Table(title="✅ 处理完成", show_header=True, header_style="bold green", box=box.ROUNDED)
        result_table.add_column("项目", style="bold")
        result_table.add_column("数值", style="cyan", justify="right")

        result_table.add_row("原始行数", str(original_count))
        result_table.add_row("去重后行数", str(dedup_count))
        result_table.add_row("删除重复行数", f"[red]{removed_count}[/red]")
        result_table.add_row("输出文件名", f"[green]{output_filename}[/green]")

        console.print(result_table)
        return True

    except Exception as e:
        import traceback
        console.print(Panel(
            f"[red]✗ 处理失败![/red]\n\n错误信息: {e}\n{traceback.format_exc()}",
            title="[bold red]错误[/bold red]",
            border_style="red"
        ))
        return False


# ────────────────────────────────────────────────────────────────────────────────
# 功能2：筛选交集（从数据文件筛选匹配行，顺序与筛选文件一致）
# ────────────────────────────────────────────────────────────────────────────────

def read_b_column(filepath, col_index_0based):
    """
    读取筛选文件指定列的所有值（从第 2 行起，跳过表头）。
    返回:
        b_order : list[str]  — 按行顺序排列的值（保留重复项用于定序，后续会去重）
        b_values: set[str]   — 用于快速查找的集合
    """
    col_1based = col_index_0based + 1
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    b_order = []   # 保留顺序（含重复）
    b_values = set()

    with Progress(
        SpinnerColumn(),
        TextColumn("[cyan]正在读取筛选文件...[/cyan]"),
        transient=True,
    ) as progress:
        progress.add_task("read_b", total=None)
        for row in ws.iter_rows(min_row=2, min_col=col_1based, max_col=col_1based):
            cell = row[0]
            if cell.value is not None:
                val = str(cell.value).strip()
                b_order.append(val)   # 记录原始顺序（含重复）
                b_values.add(val)

    wb.close()
    return b_order, b_values


def filter_and_write(file_a_path, col_a_index, b_order, b_values, output_path):
    """
    读取数据文件，将匹配行按照筛选文件的值顺序写入输出文件。
    返回: (original_rows, matched_rows)
    """

    # ── Step 1：将数据文件所有行读入内存，构建 key → [行数据, ...] 的映射 ──
    wb_a = openpyxl.load_workbook(file_a_path, read_only=True, data_only=True)
    ws_a = wb_a.active
    max_row = ws_a.max_row
    max_col = ws_a.max_column

    header_row = None
    key_to_rows: dict[str, list] = {}

    with Progress(
        SpinnerColumn(),
        TextColumn("[cyan]正在读取数据文件...[/cyan]"),
        BarColumn(),
        TaskProgressColumn(),
        transient=True,
    ) as progress:
        task = progress.add_task("read_a", total=max_row)
        for row_num, row in enumerate(ws_a.iter_rows(), start=1):

            # 表头单独保存
            if row_num == 1:
                header_row = [
                    (str(cell.value).strip() if cell.value is not None else None)
                    for cell in row
                ]
                progress.advance(task)
                continue

            key_cell = row[col_a_index]
            if key_cell.value is None:
                progress.advance(task)
                continue

            key = str(key_cell.value).strip()

            # 只缓存在 b_values 中存在的行，节省内存
            if key in b_values:
                row_data = [
                    (str(cell.value).strip() if cell.value is not None else None)
                    for cell in row
                ]
                if key not in key_to_rows:
                    key_to_rows[key] = []
                key_to_rows[key].append(row_data)

            progress.advance(task)

    wb_a.close()

    original_rows = max_row - 1  # 不含表头

    # ── Step 2：按 b_order 的顺序拼接最终输出行列表 ──
    seen_keys = set()
    ordered_rows = []

    for val in b_order:
        if val in seen_keys:
            continue
        seen_keys.add(val)
        if val in key_to_rows:
            ordered_rows.extend(key_to_rows[val])   # 同 key 多行，组内保持原顺序

    matched_rows = len(ordered_rows)

    # ── Step 3：写入输出文件 ──
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active

    all_rows = [header_row] + ordered_rows  # 表头 + 数据

    with Progress(
        SpinnerColumn(),
        TextColumn("[cyan]正在写入输出文件...[/cyan]"),
        BarColumn(),
        TaskProgressColumn(),
        transient=True,
    ) as progress:
        task = progress.add_task("write", total=len(all_rows))
        for new_row_idx, row_data in enumerate(all_rows, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                dst_cell = ws_out.cell(row=new_row_idx, column=col_idx)
                dst_cell.value = value
                dst_cell.number_format = openpyxl_numbers.FORMAT_TEXT
            progress.advance(task)

    wb_out.save(output_path)
    wb_out.close()

    return original_rows, matched_rows


def filter_intersection():
    """从数据文件中筛选匹配行，输出顺序与筛选文件保持一致"""
    try:
        console.print(Panel.fit(
            "[bold cyan]🔗 Excel 行筛选工具[/bold cyan]\n"
            "[dim]从数据文件中筛选匹配行，输出顺序与筛选文件保持一致[/dim]",
            border_style="cyan",
            padding=(1, 4),
        ))
        console.print()  # 空行美化

        xlsx_files, script_dir = get_xlsx_files()

        if not xlsx_files:
            console.print(Panel(
                "[bold red]❌ 当前文件夹中没有找到任何 .xlsx 文件！\n"
                "请将需要处理的 Excel 文件放在本程序所在文件夹中。[/bold red]",
                border_style="red",
            ))
            Prompt.ask("\n[dim]按回车返回上级菜单[/dim]", default="q")
            return False

        # ── 选择数据文件（A 文件）──
        console.rule("[bold cyan]第一步：选择数据文件（被筛选的文件）[/bold cyan]")
        display_file_list_tui(xlsx_files, script_dir)
        file_a_idx = select_file_tui(xlsx_files, prompt_text="请选择数据文件的编号或输入 'q' 返回")
        if file_a_idx is None:
            return False
        file_a_name = get_safe_filename(xlsx_files[file_a_idx])
        file_a_path = os.path.join(script_dir, file_a_name)
        console.print(f"[green]✅ 已选择数据文件：[bold white]{file_a_name}[/bold white][/green]")

        # ── 选择筛选文件（B 文件）──
        console.print()
        console.rule("[bold cyan]第二步：选择筛选文件（提供匹配值的文件）[/bold cyan]")
        display_file_list_tui(xlsx_files, script_dir)
        file_b_idx = select_file_tui(xlsx_files, prompt_text="请选择筛选文件的编号或输入 'q' 返回")
        if file_b_idx is None:
            return False
        file_b_name = get_safe_filename(xlsx_files[file_b_idx])
        file_b_path = os.path.join(script_dir, file_b_name)
        console.print(f"[green]✅ 已选择筛选文件：[bold white]{file_b_name}[/bold white][/green]")

        # ── 选择数据文件的匹配列 ──
        console.print()
        console.rule("[bold cyan]第三步：选择数据文件中用于匹配的列[/bold cyan]")
        with console.status("[cyan]正在读取数据文件列信息...[/cyan]", spinner="dots"):
            wb_a_tmp = openpyxl.load_workbook(file_a_path, read_only=True, data_only=True)
            ws_a_tmp = wb_a_tmp.active

        result_a = select_column_interactively(
            ws_a_tmp,
            prompt_title="请输入数据文件中用于匹配的列号（如 A、B、P、AA 等），输入 q 返回",
        )
        wb_a_tmp.close()
        if result_a is None:
            return False
        col_a_letter, col_a_index = result_a
        console.print(f"[green]✅ 已选择数据文件匹配列：[bold white]{col_a_letter} 列[/bold white][/green]")

        # ── 选择筛选文件的匹配列 ──
        console.print()
        console.rule("[bold cyan]第四步：选择筛选文件中提供匹配值的列[/bold cyan]")
        with console.status("[cyan]正在读取筛选文件列信息...[/cyan]", spinner="dots"):
            wb_b_tmp = openpyxl.load_workbook(file_b_path, read_only=True, data_only=True)
            ws_b_tmp = wb_b_tmp.active

        result_b = select_column_interactively(
            ws_b_tmp,
            prompt_title="请输入筛选文件中提供匹配值的列号（如 A、B、P、AA 等），输入 q 返回",
        )
        wb_b_tmp.close()
        if result_b is None:
            return False
        col_b_letter, col_b_index = result_b
        console.print(f"[green]✅ 已选择筛选文件匹配列：[bold white]{col_b_letter} 列[/bold white][/green]")

        # ── 生成输出文件名 ──
        base_name = os.path.splitext(file_a_name)[0]
        output_filename = f"{base_name}_筛选结果.xlsx"
        output_path = os.path.join(script_dir, output_filename)

        # ── 确认操作 ──
        console.print()
        console.print(Panel(
            f"[bold]操作确认[/bold]\n\n"
            f"  📄 数据文件：  [cyan]{file_a_name}[/cyan]（匹配列：{col_a_letter} 列）\n"
            f"  🔑 筛选文件：  [cyan]{file_b_name}[/cyan]（取值列：{col_b_letter} 列）\n"
            f"  🔀 输出顺序：  与筛选文件 {col_b_letter} 列顺序一致\n"
            f"  💾 输出文件：  [cyan]{output_filename}[/cyan]\n"
            f"  📁 输出目录：  [cyan]{script_dir}[/cyan]",
            border_style="yellow",
            title="[yellow]请确认[/yellow]",
        ))

        confirm = Prompt.ask(
            "[bold yellow]确认开始处理？[/bold yellow]",
            choices=["y", "n"],
            default="y",
        )
        if confirm.lower() != "y":
            console.print("[dim]已取消操作。[/dim]")
            return False

        # ── 执行处理 ──
        console.print()
        console.rule("[bold cyan]处理中[/bold cyan]")

        b_order, b_values = read_b_column(file_b_path, col_b_index)
        console.print(
            f"[green]✅ 筛选文件 {col_b_letter} 列共读取到 "
            f"[bold]{len(b_order)}[/bold] 个值（[bold]{len(b_values)}[/bold] 个唯一值）[/green]"
        )

        original_rows, matched_rows = filter_and_write(
            file_a_path, col_a_index, b_order, b_values, output_path
        )

        # ── 展示结果 ──
        console.print()
        result_table = Table(title="✅ 处理完成", show_header=True, header_style="bold green", box=box.ROUNDED)
        result_table.add_column("项目", style="bold")
        result_table.add_column("数值", style="cyan", justify="right")

        result_table.add_row("数据文件总行数（不含表头）", str(original_rows))
        result_table.add_row("筛选文件唯一值数量", str(len(b_values)))
        result_table.add_row("命中并保留的行数", f"[green]{matched_rows}[/green]")
        result_table.add_row("未命中丢弃的行数", f"[red]{original_rows - matched_rows}[/red]")
        result_table.add_row("输出顺序", f"与筛选文件 {col_b_letter} 列一致")
        result_table.add_row("输出文件名", f"[green]{output_filename}[/green]")

        console.print(result_table)
        return True

    except Exception as e:
        import traceback
        console.print(Panel(
            f"[red]✗ 处理失败![/red]\n\n错误信息: {e}\n{traceback.format_exc()}",
            title="[bold red]错误[/bold red]",
            border_style="red"
        ))
        return False


# ────────────────────────────────────────────────────────────────────────────────
# 功能3：合并当前文件夹下所有 Excel 文件
# ────────────────────────────────────────────────────────────────────────────────

def merge_all_xlsx():
    """将当前文件夹下所有 xlsx 文件纵向合并为一个文件（表头相同）"""
    try:
        console.print(Panel.fit(
            "[bold cyan]📑 Excel 批量合并工具[/bold cyan]\n"
            "[dim]将当前文件夹下所有 Excel 文件按名称合并为一个文件[/dim]",
            border_style="cyan",
            padding=(1, 4),
        ))
        console.print()

        xlsx_files, script_dir = get_xlsx_files()

        if not xlsx_files:
            console.print(Panel(
                "[bold red]❌ 当前文件夹中没有找到任何 .xlsx 文件！\n"
                "请将需要合并的 Excel 文件放在本程序所在文件夹中。[/bold red]",
                border_style="red",
            ))
            Prompt.ask("\n[dim]按回车返回上级菜单[/dim]", default="q")
            return False

        if len(xlsx_files) < 2:
            console.print(Panel(
                "[bold yellow]⚠️ 当前文件夹中只有 1 个 .xlsx 文件，无需合并。[/bold yellow]",
                border_style="yellow",
            ))
            Prompt.ask("\n[dim]按回车返回上级菜单[/dim]", default="q")
            return False

        # 显示将要合并的文件列表
        display_file_list_tui(xlsx_files, script_dir, title="📂 将要合并的 Excel 文件")

        # 让用户指定输出文件名
        console.print()
        default_output = "合并结果.xlsx"
        output_filename = Prompt.ask(
            "[bold green]请输入输出文件名[/bold green]",
            default=default_output,
        ).strip()
        if not output_filename.endswith(".xlsx"):
            output_filename += ".xlsx"
        output_path = os.path.join(script_dir, output_filename)

        # 确认操作
        console.print()
        console.print(Panel(
            f"[bold]操作确认[/bold]\n\n"
            f"  📄 合并文件数：[cyan]{len(xlsx_files)}[/cyan] 个\n"
            f"  💾 输出文件：  [cyan]{output_filename}[/cyan]\n"
            f"  📁 输出目录：  [cyan]{script_dir}[/cyan]",
            border_style="yellow",
            title="[yellow]请确认[/yellow]",
        ))

        confirm = Prompt.ask(
            "[bold yellow]确认开始合并？[/bold yellow]",
            choices=["y", "n"],
            default="y",
        )
        if confirm.lower() != "y":
            console.print("[dim]已取消操作。[/dim]")
            return False

        # 执行合并
        console.print()
        console.rule("[bold cyan]处理中[/bold cyan]")

        all_headers = None   # 以第一个文件的表头为准
        all_rows = []
        file_count = 0

        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TaskProgressColumn(),
            console=console,
        ) as progress:
            task = progress.add_task("[cyan]正在读取文件...", total=len(xlsx_files))

            for file_path in xlsx_files:
                fname = get_safe_filename(file_path)
                # 跳过输出文件本身
                if fname == output_filename:
                    progress.advance(task)
                    continue
                try:
                    headers, data_rows = _read_xlsx_to_rows(str(file_path))
                    if all_headers is None:
                        all_headers = headers
                    # 对齐列宽（多退少补）
                    width = len(all_headers)
                    for row in data_rows:
                        while len(row) < width:
                            row.append(None)
                        if len(row) > width:
                            row[:] = row[:width]
                    all_rows.extend(data_rows)
                    file_count += 1
                    progress.update(task, description=f"[cyan]已读取：{fname}（{len(data_rows)} 行）")
                except Exception as e:
                    console.print(f"[red]❌ 读取失败：{fname}，错误：{e}[/red]")
                progress.advance(task)

        if file_count == 0 or all_headers is None:
            console.print("[red]没有成功读取任何文件，合并终止。[/red]")
            return False

        with console.status("[bold cyan]正在保存文件...[/bold cyan]", spinner="dots"):
            _write_rows_to_xlsx(output_path, all_headers, all_rows)

        # 展示结果
        console.print()
        result_table = Table(title="✅ 合并完成", show_header=True, header_style="bold green", box=box.ROUNDED)
        result_table.add_column("项目", style="bold")
        result_table.add_column("数值", style="cyan", justify="right")
        result_table.add_row("合并文件数", str(file_count))
        result_table.add_row("总行数（不含表头）", str(len(all_rows)))
        result_table.add_row("总列数", str(len(all_headers)))
        result_table.add_row("输出文件名", f"[green]{output_filename}[/green]")
        console.print(result_table)
        return True

    except Exception as e:
        import traceback
        console.print(Panel(
            f"[red]✗ 处理失败![/red]\n\n错误信息: {e}\n{traceback.format_exc()}",
            title="[bold red]错误[/bold red]",
            border_style="red",
        ))
        return False


# ────────────────────────────────────────────────────────────────────────────────
# 功能4：按指定行数分割 Excel 文件
# ────────────────────────────────────────────────────────────────────────────────

def split_xlsx_by_rows():
    """将一个 Excel 文件按指定行数分割为多个子文件，每个子文件保留表头"""
    try:
        console.print(Panel.fit(
            "[bold cyan]✂️ Excel 文件分割工具[/bold cyan]\n"
            "[dim]将一个 Excel 文件按指定行数分割为多个子文件[/dim]",
            border_style="cyan",
            padding=(1, 4),
        ))
        console.print()

        xlsx_files, script_dir = get_xlsx_files()

        if not xlsx_files:
            console.print(Panel(
                "[bold red]❌ 当前文件夹中没有找到任何 .xlsx 文件！\n"
                "请将需要分割的 Excel 文件放在本程序所在文件夹中。[/bold red]",
                border_style="red",
            ))
            Prompt.ask("\n[dim]按回车返回上级菜单[/dim]", default="q")
            return False

        # 选择文件
        display_file_list_tui(xlsx_files, script_dir)
        file_idx = select_file_tui(xlsx_files)
        if file_idx is None:
            return False
        selected_file = get_safe_filename(xlsx_files[file_idx])
        filepath = os.path.join(script_dir, selected_file)
        console.print(f"[green]✅ 已选择文件：[bold white]{selected_file}[/bold white][/green]")

        # 读取文件获取行数信息
        with console.status("[bold cyan]正在读取文件...[/bold cyan]", spinner="dots"):
            headers, data_rows = _read_xlsx_to_rows(filepath)

        total_rows = len(data_rows)
        console.print(f"[green]✅ 文件共 [bold]{total_rows}[/bold] 行数据（不含表头）[/green]")

        if total_rows == 0:
            console.print("[yellow]文件没有数据行，无需分割。[/yellow]")
            return False

        # 让用户输入每份的行数
        while True:
            rows_input = Prompt.ask(
                "\n[bold green]请输入每份的行数（不含表头），输入 q 返回[/bold green]",
                default="5000",
            ).strip()

            if rows_input.lower() == "q":
                console.print("[dim]已返回上级菜单。[/dim]")
                return False

            if not rows_input.isdigit() or int(rows_input) <= 0:
                console.print("[bold red]❌ 请输入一个正整数！[/bold red]")
                continue

            rows_per_file = int(rows_input)
            break

        # 计算分割信息
        num_files = math.ceil(total_rows / rows_per_file)

        # 确认操作
        console.print()
        console.print(Panel(
            f"[bold]操作确认[/bold]\n\n"
            f"  📄 源文件：    [cyan]{selected_file}[/cyan]\n"
            f"  📊 总行数：    [cyan]{total_rows}[/cyan] 行\n"
            f"  ✂️  每份行数：  [cyan]{rows_per_file}[/cyan] 行\n"
            f"  📁 预计生成：  [cyan]{num_files}[/cyan] 个文件\n"
            f"  📁 输出目录：  [cyan]{script_dir}[/cyan]",
            border_style="yellow",
            title="[yellow]请确认[/yellow]",
        ))

        confirm = Prompt.ask(
            "[bold yellow]确认开始分割？[/bold yellow]",
            choices=["y", "n"],
            default="y",
        )
        if confirm.lower() != "y":
            console.print("[dim]已取消操作。[/dim]")
            return False

        # 执行分割
        console.print()
        console.rule("[bold cyan]处理中[/bold cyan]")

        base_name = os.path.splitext(selected_file)[0]
        output_files = []

        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TaskProgressColumn(),
            console=console,
        ) as progress:
            task = progress.add_task("[cyan]正在分割文件...", total=num_files)

            for i in range(num_files):
                start_row = i * rows_per_file
                end_row = min(start_row + rows_per_file, total_rows)
                chunk = data_rows[start_row:end_row]

                part_num = i + 1
                output_filename = f"{base_name}_第{part_num}份.xlsx"
                part_output_path = os.path.join(script_dir, output_filename)

                _write_rows_to_xlsx(part_output_path, headers, chunk)

                output_files.append((output_filename, len(chunk)))
                progress.update(task, description=f"[cyan]已写入：{output_filename}（{len(chunk)} 行）")
                progress.advance(task)

        # 展示结果
        console.print()
        result_table = Table(title="✅ 分割完成", show_header=True, header_style="bold green", box=box.ROUNDED)
        result_table.add_column("文件名", style="bold")
        result_table.add_column("行数", style="cyan", justify="right")

        for fname, row_count in output_files:
            result_table.add_row(fname, str(row_count))

        result_table.add_row("─" * 20, "─" * 6, style="dim")
        result_table.add_row("[bold]总计[/bold]", f"[bold]{total_rows}[/bold]")
        console.print(result_table)
        return True

    except Exception as e:
        import traceback
        console.print(Panel(
            f"[red]✗ 处理失败![/red]\n\n错误信息: {e}\n{traceback.format_exc()}",
            title="[bold red]错误[/bold red]",
            border_style="red",
        ))
        return False


# ────────────────────────────────────────────────────────────────────────────────
# 功能5：按主键横向合并两个 Excel 文件
# ────────────────────────────────────────────────────────────────────────────────

def merge_two_xlsx_by_key():
    """按主键列将 B 文件的列横向追加到 A 文件右侧"""
    try:
        console.print(Panel.fit(
            "[bold cyan]🔀 Excel 横向合并工具[/bold cyan]\n"
            "[dim]按主键将两个 Excel 文件横向合并（B 追加到 A 右侧）[/dim]",
            border_style="cyan",
            padding=(1, 4),
        ))
        console.print()

        xlsx_files, script_dir = get_xlsx_files()

        if not xlsx_files:
            console.print(Panel(
                "[bold red]❌ 当前文件夹中没有找到任何 .xlsx 文件！\n"
                "请将需要合并的 Excel 文件放在本程序所在文件夹中。[/bold red]",
                border_style="red",
            ))
            Prompt.ask("\n[dim]按回车返回上级菜单[/dim]", default="q")
            return False

        # ── 选择 A 文件 ──
        console.rule("[bold cyan]第一步：选择主文件（A 文件）[/bold cyan]")
        display_file_list_tui(xlsx_files, script_dir)
        file_a_idx = select_file_tui(xlsx_files, prompt_text="请选择主文件的编号或输入 'q' 返回")
        if file_a_idx is None:
            return False
        file_a_name = get_safe_filename(xlsx_files[file_a_idx])
        file_a_path = os.path.join(script_dir, file_a_name)
        console.print(f"[green]✅ 已选择主文件：[bold white]{file_a_name}[/bold white][/green]")

        # ── 选择 B 文件 ──
        console.print()
        console.rule("[bold cyan]第二步：选择追加文件（B 文件）[/bold cyan]")
        display_file_list_tui(xlsx_files, script_dir)
        file_b_idx = select_file_tui(xlsx_files, prompt_text="请选择追加文件的编号或输入 'q' 返回")
        if file_b_idx is None:
            return False
        file_b_name = get_safe_filename(xlsx_files[file_b_idx])
        file_b_path = os.path.join(script_dir, file_b_name)
        console.print(f"[green]✅ 已选择追加文件：[bold white]{file_b_name}[/bold white][/green]")

        # ── 读取两个文件 ──
        with console.status("[bold cyan]正在读取文件...[/bold cyan]", spinner="dots"):
            tbl_a = SimpleTable.from_xlsx(file_a_path)
            tbl_b = SimpleTable.from_xlsx(file_b_path)

        console.print(f"[green]✅ A 文件：{len(tbl_a)} 行 × {len(tbl_a.columns)} 列[/green]")
        console.print(f"[green]✅ B 文件：{len(tbl_b)} 行 × {len(tbl_b.columns)} 列[/green]")

        # ── 选择 A 文件的主键列 ──
        console.print()
        console.rule("[bold cyan]第三步：选择 A 文件中的主键列[/bold cyan]")
        result_a = select_column_from_df(tbl_a, prompt_title="请输入 A 文件的主键列号（如 A、B），输入 q 返回")
        if result_a is None:
            return False
        col_a_letter, col_a_index = result_a
        col_a_name = tbl_a.columns[col_a_index]
        console.print(f"[green]✅ A 文件主键列：[bold white]{col_a_letter} 列（{col_a_name}）[/bold white][/green]")

        # ── 选择 B 文件的主键列 ──
        console.print()
        console.rule("[bold cyan]第四步：选择 B 文件中的主键列[/bold cyan]")
        result_b = select_column_from_df(tbl_b, prompt_title="请输入 B 文件的主键列号（如 A、B），输入 q 返回")
        if result_b is None:
            return False
        col_b_letter, col_b_index = result_b
        col_b_name = tbl_b.columns[col_b_index]
        console.print(f"[green]✅ B 文件主键列：[bold white]{col_b_letter} 列（{col_b_name}）[/bold white][/green]")

        # ── 检查主键不匹配情况 ──
        keys_a = set(str(v).strip() for v in tbl_a[col_a_name] if v is not None)
        keys_b = set(str(v).strip() for v in tbl_b[col_b_name] if v is not None)
        only_in_a = keys_a - keys_b
        only_in_b = keys_b - keys_a

        how = "left"  # 默认以 A 为准

        if only_in_a or only_in_b:
            console.print()
            if only_in_a:
                console.print(
                    f"[yellow]⚠️  A 文件中有 [bold]{len(only_in_a)}[/bold] 个主键在 B 文件中不存在"
                    f"（B 的列将填空）[/yellow]"
                )
            if only_in_b:
                console.print(
                    f"[yellow]⚠️  B 文件中有 [bold]{len(only_in_b)}[/bold] 个主键在 A 文件中不存在[/yellow]"
                )

            console.print()
            console.print("[bold]请选择不匹配行的处理方式：[/bold]")
            console.print("  [cyan]1.[/cyan] 仅保留 A 文件的行（B 中多余的丢弃，A 中无匹配的 B 列留空）")
            console.print("  [cyan]2.[/cyan] 保留所有行（A 和 B 各自独有的行都保留，缺失列留空）")
            console.print("  [cyan]3.[/cyan] 仅保留两者都有的行（取交集）")

            merge_choice = Prompt.ask(
                "\n[bold green]请选择[/bold green]",
                choices=["1", "2", "3"],
                default="1",
            )
            if merge_choice == "2":
                how = "outer"
            elif merge_choice == "3":
                how = "inner"

        # ── 确认操作 ──
        how_desc = {"left": "以 A 文件为准", "outer": "保留所有行", "inner": "仅保留交集"}
        console.print()
        console.print(Panel(
            f"[bold]操作确认[/bold]\n\n"
            f"  📄 主文件(A)：  [cyan]{file_a_name}[/cyan]（主键列：{col_a_letter} - {col_a_name}）\n"
            f"  📄 追加文件(B)：[cyan]{file_b_name}[/cyan]（主键列：{col_b_letter} - {col_b_name}）\n"
            f"  🔀 合并方式：   [cyan]{how_desc[how]}[/cyan]\n"
            f"  📁 输出目录：   [cyan]{script_dir}[/cyan]",
            border_style="yellow",
            title="[yellow]请确认[/yellow]",
        ))

        confirm = Prompt.ask(
            "[bold yellow]确认开始合并？[/bold yellow]",
            choices=["y", "n"],
            default="y",
        )
        if confirm.lower() != "y":
            console.print("[dim]已取消操作。[/dim]")
            return False

        # ── 执行合并 ──
        console.print()
        console.rule("[bold cyan]处理中[/bold cyan]")

        with console.status("[bold cyan]正在合并数据...[/bold cyan]", spinner="dots"):
            # B 文件要追加的列（排除主键列本身）
            b_extra_indices = [i for i, c in enumerate(tbl_b.columns) if i != col_b_index]
            b_extra_headers = [tbl_b.columns[i] for i in b_extra_indices]
            # 重名处理：与 A 列名重复则加 _B 后缀
            b_extra_headers_final = []
            for h in b_extra_headers:
                if h in tbl_a.columns:
                    b_extra_headers_final.append(f"{h}_B")
                else:
                    b_extra_headers_final.append(h)

            # 构建 B 文件的主键→行映射（同一主键可能有多行，取第一行）
            b_key_map = {}
            for row in tbl_b.get_rows():
                key = str(row[col_b_index]).strip() if row[col_b_index] is not None else ""
                if key not in b_key_map:
                    b_key_map[key] = [row[i] for i in b_extra_indices]

            merged_headers = tbl_a.columns + b_extra_headers_final
            merged_rows = []
            matched_a_keys = set()

            # 遍历 A 的每一行
            for row_a in tbl_a.get_rows():
                key = str(row_a[col_a_index]).strip() if row_a[col_a_index] is not None else ""
                b_vals = b_key_map.get(key)
                if b_vals is not None:
                    matched_a_keys.add(key)
                    merged_rows.append(list(row_a) + list(b_vals))
                else:
                    if how == "inner":
                        continue  # 交集模式下跳过
                    merged_rows.append(list(row_a) + [None] * len(b_extra_indices))

            # outer 模式：追加 B 中独有的行
            if how == "outer":
                empty_a = [None] * len(tbl_a.columns)
                for row_b in tbl_b.get_rows():
                    key = str(row_b[col_b_index]).strip() if row_b[col_b_index] is not None else ""
                    if key not in matched_a_keys:
                        new_row = list(empty_a)
                        new_row[col_a_index] = row_b[col_b_index]  # 保留主键值
                        new_row += [row_b[i] for i in b_extra_indices]
                        merged_rows.append(new_row)

        # 保存
        base_name_a = os.path.splitext(file_a_name)[0]
        output_filename = f"{base_name_a}_横向合并.xlsx"
        output_path = os.path.join(script_dir, output_filename)

        with console.status("[bold cyan]正在保存文件...[/bold cyan]", spinner="dots"):
            _write_rows_to_xlsx(output_path, merged_headers, merged_rows)

        # 展示结果
        console.print()
        result_table = Table(title="✅ 横向合并完成", show_header=True, header_style="bold green", box=box.ROUNDED)
        result_table.add_column("项目", style="bold")
        result_table.add_column("数值", style="cyan", justify="right")
        result_table.add_row("A 文件行数", str(len(tbl_a)))
        result_table.add_row("B 文件行数", str(len(tbl_b)))
        result_table.add_row("合并后行数", str(len(merged_rows)))
        result_table.add_row("合并后列数", str(len(merged_headers)))
        result_table.add_row("合并方式", how_desc[how])
        result_table.add_row("输出文件名", f"[green]{output_filename}[/green]")
        console.print(result_table)
        return True

    except Exception as e:
        import traceback
        console.print(Panel(
            f"[red]✗ 处理失败![/red]\n\n错误信息: {e}\n{traceback.format_exc()}",
            title="[bold red]错误[/bold red]",
            border_style="red",
        ))
        return False


# ────────────────────────────────────────────────────────────────────────────────
# 功能6：按分组 ID 排序并标红最新时间行
# ────────────────────────────────────────────────────────────────────────────────

_COMMON_TIME_FORMATS = [
    "%Y-%m-%d %H:%M:%S",
    "%Y/%m/%d %H:%M:%S",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y/%m/%d %H:%M",
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%Y%m%d%H%M%S",
    "%Y%m%d",
]


def _detect_time_format(time_values):
    """自动检测时间列的格式，返回最佳匹配的 format 字符串；失败返回 None。
    time_values: list[str|None]"""
    samples = [str(v).strip() for v in time_values if v is not None][:100]
    if not samples:
        return None

    best_fmt = None
    best_count = 0

    for fmt in _COMMON_TIME_FORMATS:
        ok = 0
        for val in samples:
            try:
                datetime.datetime.strptime(val, fmt)
                ok += 1
            except (ValueError, TypeError):
                pass
        if ok > best_count:
            best_count = ok
            best_fmt = fmt

    # 至少 60% 匹配才算成功
    if best_fmt and best_count >= len(samples) * 0.6:
        return best_fmt
    return None


def _parse_datetime_safe(val, fmt=None):
    """安全解析时间字符串，返回 datetime 或 None"""
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    if fmt:
        try:
            return datetime.datetime.strptime(s, fmt)
        except (ValueError, TypeError):
            return None
    # 无指定格式时逐一尝试
    for f in _COMMON_TIME_FORMATS:
        try:
            return datetime.datetime.strptime(s, f)
        except (ValueError, TypeError):
            continue
    return None


def highlight_latest_rows():
    """按分组 ID 列排序，标红每组中时间最新的行"""
    try:
        console.print(Panel.fit(
            "[bold cyan]🔴 Excel 分组标红工具[/bold cyan]\n"
            "[dim]按分组 ID 列排序，标红每组中时间最新的行[/dim]",
            border_style="cyan",
            padding=(1, 4),
        ))
        console.print()

        xlsx_files, script_dir = get_xlsx_files()

        if not xlsx_files:
            console.print(Panel(
                "[bold red]❌ 当前文件夹中没有找到任何 .xlsx 文件！\n"
                "请将需要处理的 Excel 文件放在本程序所在文件夹中。[/bold red]",
                border_style="red",
            ))
            Prompt.ask("\n[dim]按回车返回上级菜单[/dim]", default="q")
            return False

        # 选择文件
        display_file_list_tui(xlsx_files, script_dir)
        file_idx = select_file_tui(xlsx_files)
        if file_idx is None:
            return False
        selected_file = get_safe_filename(xlsx_files[file_idx])
        filepath = os.path.join(script_dir, selected_file)
        console.print(f"[green]✅ 已选择文件：[bold white]{selected_file}[/bold white][/green]")

        # 读取
        with console.status("[bold cyan]正在读取文件...[/bold cyan]", spinner="dots"):
            tbl = SimpleTable.from_xlsx(filepath)

        console.print(f"[green]✅ 文件共 [bold]{len(tbl)}[/bold] 行 × [bold]{len(tbl.columns)}[/bold] 列[/green]")

        # 选择分组 ID 列
        console.print()
        console.rule("[bold cyan]选择分组 ID 列（类似 sessionId）[/bold cyan]")
        result_id = select_column_from_df(tbl, prompt_title="请输入分组 ID 列号（如 A、B），输入 q 返回")
        if result_id is None:
            return False
        id_col_letter, id_col_index = result_id
        id_col_name = tbl.columns[id_col_index]
        console.print(f"[green]✅ 分组 ID 列：[bold white]{id_col_letter} 列（{id_col_name}）[/bold white][/green]")

        # 选择时间列
        console.print()
        console.rule("[bold cyan]选择时间列[/bold cyan]")
        result_time = select_column_from_df(tbl, prompt_title="请输入时间列号（如 A、B），输入 q 返回")
        if result_time is None:
            return False
        time_col_letter, time_col_index = result_time
        time_col_name = tbl.columns[time_col_index]
        console.print(f"[green]✅ 时间列：[bold white]{time_col_letter} 列（{time_col_name}）[/bold white][/green]")

        # 自动检测时间格式
        console.print()
        time_col_values = tbl[time_col_name]
        with console.status("[bold cyan]正在检测时间格式...[/bold cyan]", spinner="dots"):
            time_fmt = _detect_time_format(time_col_values)

        if time_fmt is None:
            console.print("[yellow]⚠️  无法自动识别时间格式，将逐一尝试常见格式。[/yellow]")

        # 解析时间值
        time_parsed = []  # list[datetime|None]
        for val in time_col_values:
            time_parsed.append(_parse_datetime_safe(val, time_fmt))

        if time_fmt is not None:
            console.print(f"[green]✅ 检测到时间格式：[bold white]{time_fmt}[/bold white][/green]")

        valid_time_count = sum(1 for t in time_parsed if t is not None)
        console.print(f"[green]✅ 成功解析 [bold]{valid_time_count}[/bold] / {len(tbl)} 个时间值[/green]")

        if valid_time_count == 0:
            console.print("[red]❌ 没有任何有效的时间值，无法继续。[/red]")
            return False

        # 确认操作
        group_ids = set(str(row[id_col_index]).strip() if row[id_col_index] is not None else "" for row in tbl.get_rows())
        group_count = len(group_ids)
        console.print()
        console.print(Panel(
            f"[bold]操作确认[/bold]\n\n"
            f"  📄 处理文件：  [cyan]{selected_file}[/cyan]\n"
            f"  🔑 分组 ID 列：[cyan]{id_col_letter} 列（{id_col_name}）[/cyan]\n"
            f"  🕐 时间列：    [cyan]{time_col_letter} 列（{time_col_name}）[/cyan]\n"
            f"  📊 分组数量：  [cyan]{group_count}[/cyan] 个\n"
            f"  🔴 标记方式：  每组中时间最新的行标红背景",
            border_style="yellow",
            title="[yellow]请确认[/yellow]",
        ))

        confirm = Prompt.ask(
            "[bold yellow]确认开始处理？[/bold yellow]",
            choices=["y", "n"],
            default="y",
        )
        if confirm.lower() != "y":
            console.print("[dim]已取消操作。[/dim]")
            return False

        # ── 执行处理 ──
        console.print()
        console.rule("[bold cyan]处理中[/bold cyan]")

        # 构建带索引的行列表：(original_index, row, parsed_time, group_key)
        indexed_rows = []
        for i, row in enumerate(tbl.get_rows()):
            gk = str(row[id_col_index]).strip() if row[id_col_index] is not None else ""
            indexed_rows.append((i, row, time_parsed[i], gk))

        # 找出每个分组中时间最新的行索引
        rows_to_red = set()
        with console.status("[bold cyan]正在分析分组数据...[/bold cyan]", spinner="dots"):
            groups = {}
            for orig_i, row, t, gk in indexed_rows:
                groups.setdefault(gk, []).append((orig_i, t))
            for gk, members in groups.items():
                valid = [(idx, t) for idx, t in members if t is not None]
                if not valid:
                    continue
                max_t = max(t for _, t in valid)
                for idx, t in valid:
                    if t == max_t:
                        rows_to_red.add(idx)

        console.print(f"[green]✅ 共 [bold]{group_count}[/bold] 个分组，需标红 [bold]{len(rows_to_red)}[/bold] 行[/green]")

        # 按分组 ID + 时间排序
        with console.status("[bold cyan]正在排序...[/bold cyan]", spinner="dots"):
            # 排序键：(group_key, parsed_time_or_min, original_index)
            _dt_min = datetime.datetime.min
            sorted_rows = sorted(
                indexed_rows,
                key=lambda x: (x[3], x[2] if x[2] is not None else _dt_min, x[0])
            )
            # 建立排序后哪些位置需要标红
            sorted_red_positions = set()
            for pos, (orig_i, row, t, gk) in enumerate(sorted_rows):
                if orig_i in rows_to_red:
                    sorted_red_positions.add(pos)

        # 用 openpyxl 写入新文件
        from openpyxl.styles import PatternFill

        base_name = os.path.splitext(selected_file)[0]
        output_filename = f"{base_name}_标红最新.xlsx"
        output_path = os.path.join(script_dir, output_filename)

        headers = tbl.columns
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active

        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TaskProgressColumn(),
            console=console,
        ) as progress:
            task = progress.add_task("[cyan]正在写入文件...", total=len(sorted_rows) + 1)

            # 写入表头
            for ci, h in enumerate(headers, start=1):
                cell = ws_out.cell(row=1, column=ci, value=h)
                cell.number_format = "@"
            progress.advance(task)

            # 写入数据行
            for pos, (orig_i, row, t, gk) in enumerate(sorted_rows):
                excel_row = pos + 2  # 表头占第 1 行
                is_red = pos in sorted_red_positions
                for ci, val in enumerate(row, start=1):
                    if _is_empty_value(val):
                        cell = ws_out.cell(row=excel_row, column=ci, value=None)
                    else:
                        cleaned = sanitize_excel_cell_value(val)
                        cell = ws_out.cell(row=excel_row, column=ci, value=cleaned)
                    cell.number_format = "@"
                    if is_red:
                        cell.fill = red_fill
                progress.advance(task)

        with console.status("[bold cyan]正在保存文件...[/bold cyan]", spinner="dots"):
            wb_out.save(output_path)
            wb_out.close()

        # 展示结果
        console.print()
        result_table = Table(title="✅ 处理完成", show_header=True, header_style="bold green", box=box.ROUNDED)
        result_table.add_column("项目", style="bold")
        result_table.add_column("数值", style="cyan", justify="right")
        result_table.add_row("总行数", str(len(tbl)))
        result_table.add_row("分组数量", str(group_count))
        result_table.add_row("标红行数", f"[red]{len(rows_to_red)}[/red]")
        result_table.add_row("排序方式", f"按 {id_col_name} 分组 + {time_col_name} 排序")
        result_table.add_row("输出文件名", f"[green]{output_filename}[/green]")
        console.print(result_table)
        return True

    except Exception as e:
        import traceback
        console.print(Panel(
            f"[red]✗ 处理失败![/red]\n\n错误信息: {e}\n{traceback.format_exc()}",
            title="[bold red]错误[/bold red]",
            border_style="red",
        ))
        return False


# ────────────────────────────────────────────────────────────────────────────────
# 功能7：JSONL 解析为 Excel
# ────────────────────────────────────────────────────────────────────────────────

def _jsonl_format_value(v):
    """将 JSONL 中的单个值序列化为适合 Excel 单元格的字符串"""
    if v is None:
        return ""
    if isinstance(v, dict):
        # 键值对展开为多行 {key:value}
        return "\n".join("{" + f"{k}:{val}" + "}" for k, val in v.items())
    if isinstance(v, list):
        return json.dumps(v, ensure_ascii=False)
    return str(v)


def jsonl_to_xlsx():
    """将 JSONL 文件解析为 Excel 文件（表头自动取第一层键名）"""
    try:
        console.print(Panel.fit(
            "[bold cyan]📄 JSONL → Excel 转换工具[/bold cyan]\n"
            "[dim]将 .jsonl 文件解析为 .xlsx，表头自动识别[/dim]",
            border_style="cyan",
            padding=(1, 4),
        ))
        console.print()

        # 列出当前目录下的 jsonl 文件
        current_dir = Path.cwd()
        jsonl_files = sorted(
            [f for f in current_dir.glob("*.jsonl") if not f.name.startswith("~$")],
            key=lambda x: x.stat().st_mtime,
            reverse=True,
        )

        if not jsonl_files:
            console.print(Panel(
                "[bold red]❌ 当前文件夹中没有找到任何 .jsonl 文件！\n"
                "请将需要转换的 JSONL 文件放在当前工作目录中。[/bold red]",
                border_style="red",
            ))
            Prompt.ask("\n[dim]按回车返回上级菜单[/dim]", default="q")
            return False

        # 展示 jsonl 文件列表
        table = Table(title="📂 当前文件夹中的 JSONL 文件", show_header=True, header_style="bold cyan", box=box.ROUNDED)
        table.add_column("编号", style="bold yellow", justify="center", width=6)
        table.add_column("文件名", style="white", overflow="fold")
        table.add_column("文件大小", style="dim", justify="right", width=10)

        for idx, f in enumerate(jsonl_files, start=1):
            table.add_row(str(idx), f.name, get_file_size_str(str(f)))

        console.print(table)

        # 选择文件
        while True:
            user_input = Prompt.ask(
                "\n[bold green]请选择要转换的文件编号或输入 'q' 返回[/bold green]",
                default="q",
            ).strip()

            if user_input.lower() == "q":
                console.print("[dim]已返回上级菜单。[/dim]")
                return False

            if not user_input.isdigit():
                console.print("[bold red]❌ 请输入有效的数字编号！[/bold red]")
                continue

            idx = int(user_input)
            if idx < 1 or idx > len(jsonl_files):
                console.print(f"[bold red]❌ 编号超出范围，请输入 1 到 {len(jsonl_files)} 之间的数字！[/bold red]")
                continue
            break

        selected_file = jsonl_files[idx - 1]
        input_path = str(selected_file)
        console.print(f"[green]✅ 已选择文件：[bold white]{selected_file.name}[/bold white][/green]")

        # 默认输出文件名
        default_output = selected_file.stem + ".xlsx"
        output_filename = Prompt.ask(
            "[bold green]请输入输出文件名[/bold green]",
            default=default_output,
        ).strip()
        if not output_filename.endswith(".xlsx"):
            output_filename += ".xlsx"
        output_path = str(current_dir / output_filename)

        # 解析 JSONL
        console.print()
        console.rule("[bold cyan]处理中[/bold cyan]")

        from collections import OrderedDict

        rows = []
        all_keys = OrderedDict()

        with console.status("[bold cyan]正在解析 JSONL 文件...[/bold cyan]", spinner="dots"):
            with open(input_path, "r", encoding="utf-8") as f:
                line_no = 0
                for line in f:
                    line_no += 1
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        data = json.loads(line)
                    except json.JSONDecodeError as e:
                        console.print(f"[yellow]⚠️  第 {line_no} 行 JSON 解析失败，已跳过：{e}[/yellow]")
                        continue
                    if not isinstance(data, dict):
                        console.print(f"[yellow]⚠️  第 {line_no} 行不是 JSON 对象，已跳过[/yellow]")
                        continue
                    rows.append(data)
                    for k in data:
                        all_keys[k] = None

        if not rows:
            console.print("[red]❌ JSONL 文件中没有解析到任何有效记录。[/red]")
            return False

        headers = list(all_keys.keys())
        console.print(f"[green]✅ 解析完成：[bold]{len(rows)}[/bold] 行，[bold]{len(headers)}[/bold] 列[/green]")

        # 确认
        console.print()
        console.print(Panel(
            f"[bold]操作确认[/bold]\n\n"
            f"  📄 源文件：    [cyan]{selected_file.name}[/cyan]\n"
            f"  📊 数据量：    [cyan]{len(rows)}[/cyan] 行 × [cyan]{len(headers)}[/cyan] 列\n"
            f"  💾 输出文件：  [cyan]{output_filename}[/cyan]\n"
            f"  📁 输出目录：  [cyan]{current_dir}[/cyan]",
            border_style="yellow",
            title="[yellow]请确认[/yellow]",
        ))

        confirm = Prompt.ask(
            "[bold yellow]确认开始转换？[/bold yellow]",
            choices=["y", "n"],
            default="y",
        )
        if confirm.lower() != "y":
            console.print("[dim]已取消操作。[/dim]")
            return False

        # 用 openpyxl 写入
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        header_font = Font(name="宋体", bold=True, size=11)
        cell_font = Font(name="宋体", size=10)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # 写入表头
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=str(h))
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "@"

        # 写入数据行
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TaskProgressColumn(),
            console=console,
        ) as progress:
            task = progress.add_task("[cyan]正在写入 Excel...", total=len(rows))

            for ri, data in enumerate(rows, 2):
                for ci, h in enumerate(headers, 1):
                    val = _jsonl_format_value(data.get(h, ""))
                    # 清洗非法字符
                    val = sanitize_excel_cell_value(val)
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.font = cell_font
                    cell.border = thin_border
                    cell.number_format = "@"
                    if isinstance(val, str) and "\n" in val:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                progress.advance(task)

        # 设置行高（默认 14）
        for row_idx in range(1, len(rows) + 2):
            ws.row_dimensions[row_idx].height = 14

        # 自动列宽
        for ci, h in enumerate(headers, 1):
            max_len = len(str(h))
            for ri in range(2, min(len(rows) + 2, 202)):  # 采样前 200 行
                v = ws.cell(row=ri, column=ci).value or ""
                first_line = v.split("\n")[0] if "\n" in v else v
                max_len = max(max_len, min(len(first_line), 60))
            ws.column_dimensions[get_column_letter(ci)].width = max_len + 4

        with console.status("[bold cyan]正在保存文件...[/bold cyan]", spinner="dots"):
            wb.save(output_path)
            wb.close()

        # 展示结果
        console.print()
        result_table = Table(title="✅ 转换完成", show_header=True, header_style="bold green", box=box.ROUNDED)
        result_table.add_column("项目", style="bold")
        result_table.add_column("数值", style="cyan", justify="right")
        result_table.add_row("数据行数", str(len(rows)))
        result_table.add_row("列数", str(len(headers)))
        result_table.add_row("输出文件名", f"[green]{output_filename}[/green]")
        console.print(result_table)
        return True

    except Exception as e:
        import traceback
        console.print(Panel(
            f"[red]✗ 处理失败![/red]\n\n错误信息: {e}\n{traceback.format_exc()}",
            title="[bold red]错误[/bold red]",
            border_style="red",
        ))
        return False


def show_banner():
    """显示欢迎横幅"""
    banner = """
    ╔═══════════════════════════════════════════════════════╗
    ║                                                       ║
    ║     📊 Excel 通用数据处理工具                         ║
    ║                                                       ║
    ║     数据提取 / 去重 / 筛选交集                        ║
    ║     批量合并 / 文件分割 / 横向合并                    ║
    ║     分组排序标红 / JSONL转Excel                       ║
    ║                                                       ║
    ╚═══════════════════════════════════════════════════════╝
    """
    console.print(banner, style="bold cyan")

def list_excel_files():
    """列出当前工作目录下的Excel文件"""
    # 使用当前工作目录而不是脚本所在目录
    current_dir = Path.cwd()
    excel_files = list(current_dir.glob('*.xlsx'))
    excel_files = [f for f in excel_files if not f.name.startswith('~$')]
    return sorted(excel_files, key=lambda x: x.stat().st_mtime, reverse=True)

def show_file_list(files):
    """显示文件列表"""
    if not files:
        console.print("[yellow]⚠ 当前目录下没有找到Excel文件[/yellow]")
        console.print(f"[dim]当前工作目录: {Path.cwd()}[/dim]")
        return

    # 获取终端宽度
    try:
        terminal_width = shutil.get_terminal_size().columns
    except (ValueError, OSError):
        terminal_width = 80
    
    # 计算文件名列的宽度
    col_widths = {'序号': 8, '大小': 12, '修改时间': 20, '间隔': 6}
    filename_width = terminal_width - sum(col_widths.values())
    filename_width = max(filename_width, 30)

    table = Table(title="可用的Excel文件", box=box.ROUNDED)
    table.add_column("序号", style="cyan", justify="center", width=8)
    table.add_column("文件名", style="green", width=filename_width, overflow="fold")
    table.add_column("大小", style="yellow", justify="right", width=12)
    table.add_column("修改时间", style="magenta", width=20)

    for idx, file in enumerate(files, 1):
        size = file.stat().st_size
        size_str = f"{size / 1024:.2f} KB" if size < 1024 * 1024 else f"{size / (1024 * 1024):.2f} MB"

        mtime = datetime.datetime.fromtimestamp(file.stat().st_mtime)
        mtime_str = mtime.strftime("%Y-%m-%d %H:%M:%S")

        # 只显示文件名
        display_name = get_safe_filename(file)
        if len(display_name) > filename_width:
            display_name = wrap_text(display_name, filename_width)

        table.add_row(str(idx), display_name, size_str, mtime_str)

    console.print(table)

def _run_data_extraction():
    """运行通用数据提取功能（功能1的子循环）"""
    console.print("\n[bold cyan]✨ 已进入 Excel 通用数据提取功能[/bold cyan]")
    while True:
        console.clear()
        show_banner()
        console.print(f"[dim]当前工作目录: {Path.cwd()}[/dim]")

        # 先显示文件列表
        excel_files = list_excel_files()
        show_file_list(excel_files)

        console.print("\n[bold cyan]╔═══════════════════════════════════════╗[/bold cyan]")
        console.print("[bold cyan]║[/bold cyan] [bold]文件选择菜单[/bold]                          [bold cyan]║[/bold cyan]")
        console.print("[bold cyan]╠═══════════════════════════════════════╣[/bold cyan]")
        console.print("[bold cyan]║[/bold cyan] [cyan]1.[/cyan] 选择文件编号                      [bold cyan]║[/bold cyan]")
        console.print("[bold cyan]║[/bold cyan] [cyan]2.[/cyan] 手动输入文件路径                  [bold cyan]║[/bold cyan]")
        console.print("[bold cyan]║[/bold cyan] [cyan]3.[/cyan] 刷新文件列表                      [bold cyan]║[/bold cyan]")
        console.print("[bold cyan]║[/bold cyan] [cyan]4.[/cyan] 返回主菜单                        [bold cyan]║[/bold cyan]")
        console.print("[bold cyan]╚═══════════════════════════════════════╝[/bold cyan]")

        choice = Prompt.ask("\n[bold green]请输入选项 [1/2/3/4][/bold green]", choices=["1", "2", "3", "4"], default="1")

        if choice == "4":
            console.print("\n[yellow]返回主菜单[/yellow]")
            return

        if choice == "3":
            continue

        file_path = None

        if choice == "1":
            if not excel_files:
                console.print("[red]✗ 没有可选择的文件[/red]")
                continue

            file_idx = Prompt.ask(
                f"请输入文件编号 (1-{len(excel_files)})",
                default="1"
            )

            try:
                idx = int(file_idx) - 1
                if 0 <= idx < len(excel_files):
                    file_path = str(excel_files[idx])
                else:
                    console.print("[red]✗ 无效的文件编号[/red]")
                    continue
            except ValueError:
                console.print("[red]✗ 请输入有效的数字[/red]")
                continue

        elif choice == "2":
            file_path = Prompt.ask("请输入文件路径")
            if not os.path.isabs(file_path):
                file_path = str(Path.cwd() / file_path)
            if not os.path.exists(file_path):
                console.print("[red]✗ 文件不存在[/red]")
                continue

        console.print(f"\n[bold cyan]开始处理文件...[/bold cyan]")
        success, result, row_count = process_excel_interactive(file_path)

        if success:
            panel = Panel(
                f"[green]✓ 处理成功![/green]\n\n"
                f"📁 输出文件: [cyan]{result}[/cyan]\n"
                f"📊 处理行数: [yellow]{row_count}[/yellow] 行",
                title="[bold green]处理完成[/bold green]",
                border_style="green"
            )
            console.print(panel)
        else:
            if result != "用户取消操作" and result != "未进行修改":
                panel = Panel(
                    f"[red]✗ 处理失败![/red]\n\n"
                    f"错误信息: {result}",
                    title="[bold red]错误[/bold red]",
                    border_style="red"
                )
                console.print(panel)


def main():
    """主函数"""
    console.clear()
    show_banner()

    # 显示当前工作目录
    console.print(f"[dim]📁 当前工作目录: {Path.cwd()}[/dim]\n")

    all_choices = ["1", "2", "3", "4", "5", "6", "7", "8", "9"]

    while True:
        # ── 功能选择菜单 ──
        console.print()
        menu_table = Table(box=box.ROUNDED, show_header=False, padding=(0, 2))
        menu_table.add_column("选项", style="cyan", width=6)
        menu_table.add_column("说明", style="white")
        menu_table.add_row("1.", "Excel 通用数据提取（JSON/键值对/纯文本）")
        menu_table.add_row("2.", "去除指定列的重复项")
        menu_table.add_row("3.", "筛选交集（按另一文件顺序筛选）")
        menu_table.add_row("4.", "合并当前文件夹下所有 Excel 文件")
        menu_table.add_row("5.", "按指定行数分割 Excel 文件")
        menu_table.add_row("6.", "按主键横向合并两个 Excel 文件")
        menu_table.add_row("7.", "分组排序并标红最新时间行")
        menu_table.add_row("8.", "JSONL 解析为 Excel 文件")
        menu_table.add_row("9.", "退出程序")
        console.print(menu_table)

        func_choice = Prompt.ask(
            "\n[bold green]请输入选项 [1-9][/bold green]",
            choices=all_choices,
            default="1",
        )

        if func_choice == "9":
            console.print("\n[yellow]👋 感谢使用,再见![/yellow]")
            return

        if func_choice == "1":
            _run_data_extraction()

        elif func_choice == "2":
            console.clear()
            deduplicate_by_column()

        elif func_choice == "3":
            console.clear()
            filter_intersection()

        elif func_choice == "4":
            console.clear()
            merge_all_xlsx()

        elif func_choice == "5":
            console.clear()
            split_xlsx_by_rows()

        elif func_choice == "6":
            console.clear()
            merge_two_xlsx_by_key()

        elif func_choice == "7":
            console.clear()
            highlight_latest_rows()

        elif func_choice == "8":
            console.clear()
            jsonl_to_xlsx()

        # 每次功能执行完毕后自动回到主菜单循环顶部
        console.clear()
        show_banner()
        console.print(f"[dim]📁 当前工作目录: {Path.cwd()}[/dim]\n")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        console.print("\n\n[yellow]程序已被用户中断[/yellow]")
    except Exception as e:
        console.print(f"\n[red]发生错误: {e}[/red]")
        import traceback
        console.print(traceback.format_exc())